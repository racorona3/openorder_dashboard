#!/usr/bin/env python3
"""
Bo - Boeing Distribution Contract Lookup Tool
Generator Script: produces a self-contained index.html for GitLab Pages

Usage:
    python generate_bo.py

Requirements:
    pip install openpyxl
"""

from openpyxl import load_workbook
import json, base64, os


# ── Paths ────────────────────────────────────────────────────────────────────

XLSX_PATH   = r"C:\Users\zn424f\OneDrive - The Boeing Company\Boeing App Projects\Project Bo\honeywell_global_consolidated_agreements.xlsx"
LOGO_PATH   = r"C:\Users\zn424f\OneDrive - The Boeing Company\Boeing App Projects\Project Bo\boeing_logo.png"
OUTPUT_PATH = r"C:\Users\zn424f\OneDrive - The Boeing Company\Boeing App Projects\Project Bo\index.html"


# ── Data Loading ─────────────────────────────────────────────────────────────

def load_data(xlsx_path):
    print(f"Loading: {xlsx_path}")
    wb = load_workbook(xlsx_path, read_only=True)

    sa_priced = {}
    ws = wb["2025 SA Priced"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        pn, price, uom, leadtime = row[0], row[1], row[2], row[3]
        if pn:
            key = str(pn).strip().upper()
            sa_priced[key] = {
                "price":    float(price) if price else 0.0,
                "uom":      str(uom) if uom else "EA",
                "leadtime": str(leadtime) if leadtime else "TBD"
            }

    ga_priced = {}
    ws = wb["2025 GA Priced"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        pn, price, uom, leadtime, category = row[0], row[1], row[2], row[3], row[4]
        if pn:
            key = str(pn).strip().upper()
            ga_priced[key] = {
                "price":    float(price) if price else 0.0,
                "uom":      str(uom) if uom else "EA",
                "leadtime": str(leadtime) if leadtime else "TBD",
                "category": str(category) if category else ""
            }

    unpriced = {}
    ws = wb["2025 Unpriced Master"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        pn = row[0]
        if pn:
            key = str(pn).strip().upper()
            unpriced[key] = {
                "amendment":   str(row[5]) if row[5] else "",
                "bd_comments": str(row[3]) if row[3] else ""
            }

    return sa_priced, ga_priced, unpriced


def logo_b64(path):
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode()


# ── HTML Builder ─────────────────────────────────────────────────────────────

def build_html(sa_priced, ga_priced, unpriced, logo_b64_str, stats):
    data_js = json.dumps(
        {"sa": sa_priced, "ga": ga_priced, "unpriced": unpriced},
        separators=(",", ":")
    )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Bo \u2014 Boeing Distribution Contract Lookup</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Rajdhani:wght@400;500;600;700&family=Share+Tech+Mono&family=DM+Sans:wght@300;400;500&display=swap" rel="stylesheet">
<style>
  :root {{
    --bg:#090d14;--surface:#0f1825;--surface2:#111c2e;--border:#2a4570;
    --accent:#1565c0;--accent2:#1e88e5;--glow:#1e88e540;
    --text:#cdd8e8;--muted:#7a9abf;
    --green:#00e676;--green-dim:#00e67620;
    --amber:#ffab00;--amber-dim:#ffab0020;
    --red:#ff5252;--red-dim:#ff525220;
    --purple:#ce93d8;--purple-dim:#ce93d815;
    --mono:'Share Tech Mono',monospace;
    --head:'Rajdhani',sans-serif;
    --body:'DM Sans',sans-serif;
  }}
  *,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
  body{{background:var(--bg);color:var(--text);font-family:var(--body);font-size:16px;min-height:100vh;overflow-x:hidden}}
  body::before{{content:'';position:fixed;inset:0;background-image:linear-gradient(var(--border) 1px,transparent 1px),linear-gradient(90deg,var(--border) 1px,transparent 1px);background-size:40px 40px;opacity:.35;pointer-events:none;z-index:0}}
  body::after{{content:'';position:fixed;top:-200px;left:50%;transform:translateX(-50%);width:900px;height:500px;background:radial-gradient(ellipse,#1565c030 0%,transparent 70%);pointer-events:none;z-index:0}}
  .wrap{{position:relative;z-index:1;max-width:900px;margin:0 auto;padding:0 24px 80px}}
  header{{display:flex;align-items:center;justify-content:space-between;padding:32px 0 28px;border-bottom:1px solid var(--border);margin-bottom:40px}}
  .brand{{display:flex;align-items:center;gap:4px}}
  .brand-b{{font-family:var(--head);font-size:52px;font-weight:700;color:#fff;line-height:1;letter-spacing:-2px}}
  .brand-logo{{width:46px;height:46px;object-fit:contain;filter:brightness(1.15) saturate(1.2);margin-bottom:2px}}
  .brand-sub{{font-family:var(--mono);font-size:11px;color:var(--muted);letter-spacing:.15em;margin-top:4px}}
  .stats-bar{{display:flex;gap:20px}}
  .stat-pill{{display:flex;flex-direction:column;align-items:flex-end;gap:2px}}
  .stat-pill .num{{font-family:var(--mono);font-size:18px;font-weight:600;color:#fff}}
  .stat-pill .lbl{{font-family:var(--mono);font-size:9px;letter-spacing:.1em;color:var(--muted);text-transform:uppercase}}
  .stat-pill.sa .num{{color:var(--green)}}.stat-pill.ga .num{{color:var(--purple)}}.stat-pill.un .num{{color:var(--amber)}}
  .search-section{{margin-bottom:32px}}
  .search-label{{font-family:var(--mono);font-size:13px;letter-spacing:.12em;color:var(--muted);text-transform:uppercase;margin-bottom:10px;display:block}}
  .search-row{{display:flex;gap:10px;align-items:flex-start}}
  textarea{{flex:1;background:var(--surface);border:1px solid var(--border);border-radius:6px;color:var(--text);font-family:var(--mono);font-size:13px;padding:14px 16px;resize:vertical;min-height:56px;transition:border-color .2s,box-shadow .2s;outline:none;line-height:1.6}}
  textarea::placeholder{{color:var(--muted)}}
  textarea:focus{{border-color:var(--accent2);box-shadow:0 0 0 3px var(--glow)}}
  .btn-lookup{{background:var(--accent);border:1px solid var(--accent2);color:#fff;font-family:var(--head);font-size:15px;font-weight:600;letter-spacing:.08em;padding:14px 28px;border-radius:6px;cursor:pointer;transition:background .15s,box-shadow .15s,transform .1s;white-space:nowrap;height:56px;display:flex;align-items:center;gap:8px}}
  .btn-lookup:hover{{background:var(--accent2);box-shadow:0 0 20px var(--glow)}}
  .btn-lookup:active{{transform:scale(.98)}}
  .hint{{font-size:13px;color:var(--muted);margin-top:8px;font-family:var(--mono)}}
  #results{{margin-top:8px}}
  .result-card{{background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:20px 24px;margin-bottom:12px;position:relative;overflow:hidden;animation:fadeSlide .2s ease both}}
  @keyframes fadeSlide{{from{{opacity:0;transform:translateY(8px)}}to{{opacity:1;transform:translateY(0)}}}}
  .result-card::before{{content:'';position:absolute;left:0;top:0;bottom:0;width:4px;border-radius:8px 0 0 8px}}
  .result-card.sa-priced{{border-color:#00e67655}}.result-card.sa-priced::before{{background:var(--green)}}
  .result-card.ga-priced{{border-color:#ce93d855}}.result-card.ga-priced::before{{background:var(--purple)}}
  .result-card.unpriced{{border-color:#ffab0055}}.result-card.unpriced::before{{background:var(--amber)}}
  .result-card.off-contract{{border-color:#ff525255}}.result-card.off-contract::before{{background:var(--red)}}
  .rc-header{{display:flex;align-items:center;justify-content:space-between;margin-bottom:14px;flex-wrap:wrap;gap:8px}}
  .rc-pn{{font-family:var(--mono);font-size:24px;font-weight:600;color:#fff;letter-spacing:.04em}}
  .badge{{font-family:var(--mono);font-size:11px;letter-spacing:.1em;padding:5px 12px;border-radius:20px;text-transform:uppercase;font-weight:600}}
  .badge.sa{{background:var(--green-dim);color:var(--green);border:1px solid #00e67650}}
  .badge.ga{{background:var(--purple-dim);color:var(--purple);border:1px solid #ce93d850}}
  .badge.un{{background:var(--amber-dim);color:var(--amber);border:1px solid #ffab0050}}
  .badge.off{{background:var(--red-dim);color:var(--red);border:1px solid #ff525250}}
  .rc-body{{display:grid;grid-template-columns:repeat(auto-fill,minmax(160px,1fr));gap:12px 20px}}
  .rc-field{{display:flex;flex-direction:column;gap:3px}}
  .rc-field .fk{{font-family:var(--mono);font-size:11px;letter-spacing:.12em;text-transform:uppercase;color:var(--muted)}}
  .rc-field .fv{{font-family:var(--mono);font-size:16px;color:#e8f0fe}}
  .rc-field .fv.price{{font-size:26px;font-weight:700;color:var(--green)}}
  .rc-field .fv.amber{{color:var(--amber)}}.rc-field .fv.muted{{color:var(--muted);font-style:italic}}
  .rc-message{{font-size:15px;color:var(--text);line-height:1.7;margin-top:4px}}
  .rc-message.off{{color:var(--red)}}
  .summary-bar{{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:16px;font-family:var(--mono);font-size:11px;color:var(--muted)}}
  .summary-bar span{{letter-spacing:.05em}}
  ::-webkit-scrollbar{{width:6px;height:6px}}::-webkit-scrollbar-track{{background:var(--surface)}}::-webkit-scrollbar-thumb{{background:var(--border);border-radius:3px}}
  @media(max-width:600px){{.stats-bar{{display:none}}.rc-body{{grid-template-columns:1fr 1fr}}.brand-b{{font-size:40px}}.brand-logo{{width:36px;height:36px}}}}
</style>
</head>
<body>
<div class="wrap">
  <header>
    <div>
      <div class="brand">
        <span class="brand-b">B</span>
        <img class="brand-logo" src="data:image/png;base64,{logo_b64_str}" alt="Boeing wingmark">
      </div>
      <div class="brand-sub">BOEING DISTRIBUTION &middot; CONTRACT LOOKUP &middot; 2025</div>
    </div>
    <div class="stats-bar">
      <div class="stat-pill sa"><span class="num">{stats['sa_priced']:,}</span><span class="lbl">SA Priced</span></div>
      <div class="stat-pill ga"><span class="num">{stats['ga_priced']:,}</span><span class="lbl">GA Priced</span></div>
      <div class="stat-pill un"><span class="num">{stats['unpriced']:,}</span><span class="lbl">Unpriced</span></div>
    </div>
  </header>
  <div class="search-section">
    <span class="search-label">&#x25B6;&nbsp; Part Number Lookup</span>
    <div class="search-row">
      <textarea id="pn-input" placeholder="Enter one or more part numbers \u2014 one per line, or comma/space separated" rows="2"></textarea>
      <button class="btn-lookup" onclick="lookup()">
        <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><circle cx="11" cy="11" r="8"/><line x1="21" y1="21" x2="16.65" y2="16.65"/></svg>
        LOOKUP
      </button>
    </div>
    <div class="hint">Accepts single part numbers, bulk paste, CSV, or line-separated lists &mdash; up to 500 at once.</div>
  </div>
  <div id="results"></div>
</div>
<script>
const DB={data_js};
function parseParts(raw){{return[...new Set(raw.split(/[\\n,;\\s]+/).map(p=>p.trim().toUpperCase()).filter(p=>p.length>0))];}}
function fmt_price(price,uom){{if(!price&&price!==0)return'\u2014';const p=Number(price);return'$'+p.toLocaleString('en-US',{{minimumFractionDigits:2,maximumFractionDigits:4}})+' / '+(uom||'EA');}}
function fmt_lead(lt){{if(!lt||lt==='None'||lt==='TBD')return'TBD';const s=String(lt);if(/^\\d+$/.test(s))return s+' wks';return s;}}
function buildCard(pn){{
  const sa=DB.sa[pn],ga=DB.ga[pn],un=DB.unpriced[pn];
  let cls,bCls,bTxt,body;
  if(sa){{
    cls='sa-priced';bCls='sa';bTxt='ON-CONTRACT \u00b7 SA PRICED';
    body=`<div class="rc-field"><span class="fk">2025 Price</span><span class="fv price">${{fmt_price(sa.price,sa.uom)}}</span></div>
    <div class="rc-field"><span class="fk">Lead Time</span><span class="fv">${{fmt_lead(sa.leadtime)}}</span></div>
    <div class="rc-field"><span class="fk">Contract</span><span class="fv">Supply Agreement</span></div>
    <div class="rc-field"><span class="fk">Status</span><span class="fv" style="color:var(--green)">Priced &amp; Active</span></div>`;
  }}else if(ga){{
    cls='ga-priced';bCls='ga';bTxt='ON-CONTRACT \u00b7 GROWTH AGREEMENT';
    body=`<div class="rc-field"><span class="fk">2025 Price</span><span class="fv price" style="color:var(--purple)">${{fmt_price(ga.price,ga.uom)}}</span></div>
    <div class="rc-field"><span class="fk">Lead Time</span><span class="fv">${{fmt_lead(ga.leadtime)}}</span></div>
    <div class="rc-field"><span class="fk">Contract</span><span class="fv">Growth Agreement</span></div>
    <div class="rc-field"><span class="fk">GA Category</span><span class="fv" style="color:var(--purple)">${{ga.category||'\u2014'}}</span></div>`;
  }}else if(un){{
    cls='unpriced';bCls='un';bTxt='ON-CONTRACT \u00b7 UNPRICED';
    body=`<div class="rc-field"><span class="fk">Price</span><span class="fv amber">Unpriced \u2014 Needs RFQ</span></div>
    <div class="rc-field"><span class="fk">Amendment</span><span class="fv">${{un.amendment||'\u2014'}}</span></div>
    <div class="rc-field"><span class="fk">BD Notes</span><span class="fv muted">${{un.bd_comments||'\u2014'}}</span></div>
    <div class="rc-field"><span class="fk">Action</span><span class="fv amber">Honeywell needs to submit RFQ to Boeing</span></div>`;
  }}else{{
    cls='off-contract';bCls='off';bTxt='OFF-CONTRACT';
    body=`<p class="rc-message off">This part number is not found on any active Honeywell Supply Agreement or Growth Agreement.<br><br>Please reach out to your <strong>Boeing Distribution order book manager</strong> or <strong>program manager</strong> to evaluate sourcing options.</p>`;
  }}
  return`<div class="result-card ${{cls}}"><div class="rc-header"><span class="rc-pn">${{pn}}</span><span class="badge ${{bCls}}">${{bTxt}}</span></div><div class="rc-body">${{body}}</div></div>`;
}}
function lookup(){{
  const raw=document.getElementById('pn-input').value;
  const parts=parseParts(raw);
  if(!parts.length)return;
  const container=document.getElementById('results');
  const counts={{sa:0,ga:0,un:0,off:0}};
  const cards=parts.map(pn=>{{
    if(DB.sa[pn])counts.sa++;else if(DB.ga[pn])counts.ga++;else if(DB.unpriced[pn])counts.un++;else counts.off++;
    return buildCard(pn);
  }});
  let summaryParts=[];
  if(parts.length>1){{
    if(counts.sa)summaryParts.push(counts.sa+' SA Priced');
    if(counts.ga)summaryParts.push(counts.ga+' Growth Agreement');
    if(counts.un)summaryParts.push(counts.un+' Unpriced');
    if(counts.off)summaryParts.push(counts.off+' Off-Contract');
    container.innerHTML=`<div class="summary-bar"><span>RESULTS: ${{parts.length}} PART(S) \u2014 </span><span>${{summaryParts.join(' \u00b7 ')}}</span></div>`+cards.join('');
  }}else{{container.innerHTML=cards.join('');}}
}}
document.getElementById('pn-input').addEventListener('keydown',function(e){{if(e.key==='Enter'&&!e.shiftKey&&!this.value.includes('\\n')){{e.preventDefault();lookup();}}}});
</script>
</body>
</html>"""


# ── Main ─────────────────────────────────────────────────────────────────────

def generate(xlsx_path, logo_path, output_path):
    sa_priced, ga_priced, unpriced = load_data(xlsx_path)
    logo = logo_b64(logo_path)

    stats = {
        "sa_priced": len(sa_priced),
        "ga_priced": len(ga_priced),
        "unpriced":  len(unpriced),
        "total":     len(sa_priced) + len(ga_priced) + len(unpriced)
    }

    html = build_html(sa_priced, ga_priced, unpriced, logo, stats)

    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"\nBo generated successfully!")
    print(f"  Output:     {output_path}")
    print(f"  SA Priced:  {stats['sa_priced']:,}")
    print(f"  GA Priced:  {stats['ga_priced']:,}")
    print(f"  Unpriced:   {stats['unpriced']:,}")
    print(f"  Total:      {stats['total']:,} parts indexed")


if __name__ == "__main__":
    generate(
        xlsx_path=XLSX_PATH,
        logo_path=LOGO_PATH,
        output_path=OUTPUT_PATH
    )
